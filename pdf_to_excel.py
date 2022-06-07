import pdfplumber
import re
from openpyxl import Workbook


# Work in progress
def vernon_ny_parser(pdf_file):
    """
    pdf_file: Specify PDF path
    """
    row = 2  # First row reserved for column_list
    address_abb = [" ST", " AVE", " PL", " AV", " PLACE", " AVENUE", " RD", " ROAD", " DR", " DRIVE", " COURT",
                   " STREET", " PLAZA", " TERRACE", " LA", " LANE", " LN", " PKY", " PARKWAY", " PKWAY", " PO BOX",
                   " NY"]
    column_list = ["Tax Map Parcel ID", "Account Number", "Property Address", "Apartment", "Number of Units",
                   "Current Owners Name", "Current Owners Address", "Acerage", "Full Market Value"]
    owner_ending_list = [", ", "LLC", "INC", "LTD", "CORP", "ASSOCIATION", "C/O"]

    wb = Workbook()
    ws = wb.active
    ws.append(column_list)
    with pdfplumber.open(pdf_file) as pdf:
        pdf_size = int(str(pdf.pages[-1])[6:-1]) - 1
        pdf_count = 0

        while pdf_count < pdf_size:
            page = pdf.pages[pdf_count]
            text = page.extract_text()
            property_blocks = text.split(
                "**********************************************************************************************")

            for property_block in property_blocks[1:-1]:
                property_block = re.split(r'\s{4,}', property_block)
                property_block = property_block[1:]

                for index, text_line in enumerate(property_block):
                    if index == 0:  # Property Address
                        cell_ref = ws.cell(row=row, column=3)
                        cell_ref.value = text_line

                    elif "ACCT: " in text_line:  # Account Number
                        cell_ref = ws.cell(row=row, column=2)
                        text_line = text_line.split(" ")[1]
                        cell_ref.value = text_line

                    elif index == 2 or "." in text_line and "-" in text_line:  # Parcel ID
                        cell_ref = ws.cell(row=row, column=1)
                        cell_ref.value = text_line

                    elif index == 3 and "   " in text_line:  # Apartment Number
                        cell_ref = ws.cell(row=row, column=4)
                        text_line = text_line.split("   ")[1]
                        cell_ref.value = text_line

                    elif any(owner in text_line for owner in owner_ending_list):  # Current Owners Name
                        cell_ref = ws.cell(row=row, column=6)
                        if cell_ref.value:
                            cell_ref.value = (cell_ref.value + "/" + text_line)
                        else:
                            cell_ref.value = text_line

                    elif any(abb in text_line for abb in address_abb):  # Owner Address
                        cell_ref = ws.cell(row=row, column=7)
                        if cell_ref.value:
                            cell_ref.value = (cell_ref.value + ", " + text_line)
                        else:
                            cell_ref.value = text_line

                    elif "UNITS" in text_line:  # Number of Units
                        cell_ref = ws.cell(row=row, column=5)
                        text_line = text_line.split(" ")[0]
                        cell_ref.value = text_line

                    elif "ACREAGE " in text_line:  # Acres
                        cell_ref = ws.cell(row=row, column=8)
                        text_line = text_line.split("  ")[1]
                        cell_ref.value = text_line

                    elif "FULL MKT VAL" in text_line:  # Full Market Value
                        cell_ref = ws.cell(row=row, column=9)
                        text_line = text_line.split("  ")[1]
                        cell_ref.value = text_line
                    else:
                        continue
                row += 1
            pdf_count = pdf_count + 1

    wb.save("vernon_real_estate_2021.xlsx")
    pdf.close()
    exit()


vernon_ny_parser(pdf_file="2021_Final_Roll_Website_Updated.pdf")
